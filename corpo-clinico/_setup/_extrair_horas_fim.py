"""
Lê os merges (células mescladas) do Mapa de Sala Excel pra descobrir o
horário de FIM de cada turno — o Excel pinta a célula do início ao fim,
e essa extensão é o período de atendimento.

Gera 05_horas_fim.sql com UPDATEs em cc_escala.
"""
import openpyxl, os, sys

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
PLAN = r'G:\Drives compartilhados\Endovascular SP\3. Geral\Mapa de Sala - Endovascular SP.xlsx'
OUT = os.path.join(os.path.dirname(__file__), '05_horas_fim.sql')

wb = openpyxl.load_workbook(PLAN, data_only=True)
ws = wb['Mapa']

# linha -> hora (coluna B)
linha_hora = {}
for r in range(4, 130):
    v = ws.cell(row=r, column=2).value
    if hasattr(v, 'strftime'): linha_hora[r] = v.strftime('%H:%M')
    elif v: linha_hora[r] = str(v)

def dia_da_linha(L):
    if 5 <= L <= 25:  return 1
    if 27 <= L <= 47: return 2
    if 49 <= L <= 69: return 3
    if 71 <= L <= 91: return 4
    if 93 <= L <= 113: return 5
    return None

# salas no Excel estão nas colunas 3..12 → sala_id 1..10 (mesma ordem do seed)
linhas = []
for m in sorted(ws.merged_cells.ranges, key=lambda m:(m.min_row, m.min_col)):
    if not (3 <= m.min_col <= 12): continue
    topo = ws.cell(row=m.min_row, column=m.min_col).value
    if not topo or not str(topo).strip(): continue
    dia = dia_da_linha(m.min_row)
    if dia is None: continue
    sala_id = m.min_col - 2
    hi = linha_hora.get(m.min_row)
    hf = linha_hora.get(m.max_row)   # hora do último slot pintado
    if not hi or not hf: continue
    linhas.append((dia, sala_id, hi, hf, str(topo)[:35]))

L = ['-- Horários de fim dos turnos — extraídos dos merges do Mapa de Sala Excel.',
     '-- Cada turno: do hora_inicio até o último slot que o Excel pinta.',
     '-- Executar no Supabase SQL Editor.', '']
for dia, sala_id, hi, hf, quem in linhas:
    L.append(f"UPDATE cc_escala SET hora_fim = '{hf}' "
             f"WHERE dia_semana = {dia} AND sala_id = {sala_id} AND hora_inicio = '{hi}';"
             f"  -- {quem}")

open(OUT, 'w', encoding='utf-8', newline='').write('\n'.join(L) + '\n')
print(f'{len(linhas)} turnos com horário de fim. SQL: {OUT}')

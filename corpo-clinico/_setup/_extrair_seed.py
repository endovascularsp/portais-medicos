"""
Extrai dados das 2 planilhas (Corpo clínico + Mapa de Sala) e gera o
seed SQL do Portal Corpo Clínico (02_seed.sql).

- cc_salas: 10 salas do Mapa
- cc_profissionais: 17 do Corpo Clínico + os que só aparecem no Mapa
- cc_escala: a grade dia×sala×profissional

IDs explícitos (1..N) pra a escala referenciar; sequences resetadas no fim.
"""
import openpyxl, re, sys, os, unicodedata

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

def norm(s):
    """minúsculo sem acento — pra comparar nomes entre as planilhas."""
    s = unicodedata.normalize('NFD', str(s or ''))
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower().strip()

PLAN_GERAL = r'G:\Drives compartilhados\Endovascular SP\3. Geral\Planilha Geral - Endovascular SP.xlsx'
PLAN_MAPA  = r'G:\Drives compartilhados\Endovascular SP\3. Geral\Mapa de Sala - Endovascular SP.xlsx'
OUT = os.path.join(os.path.dirname(__file__), '02_seed.sql')

def esc(v):
    if v is None: return 'NULL'
    return "'" + str(v).strip().replace("'", "''") + "'"

def num(v):
    if v is None or v == '': return 'NULL'
    try: return str(float(v))
    except: return 'NULL'

def boolsql(v):
    return 'TRUE' if str(v).strip().lower() in ('sim', 'true', '1') else 'FALSE'

def limpa_tel(v):
    if v is None: return None
    s = str(v).strip()
    if s.endswith('.0'): s = s[:-2]
    return s

def conselho_de(espec):
    e = (espec or '').lower()
    if 'fisio' in e: return 'CREFITO'
    if 'nutri' in e: return 'CRN'
    return 'CRM'

# ===== 1. CORPO CLÍNICO =====
wb = openpyxl.load_workbook(PLAN_GERAL, read_only=True, data_only=True)
ws = wb['Corpo clínico']
profs = []
for r in ws.iter_rows(min_row=4, values_only=True):
    if not r or not r[2]:  # col Nome (índice 2)
        continue
    profs.append({
        'nome': str(r[2]).strip(),
        'especialidade': r[3],
        'observacoes': r[4],            # "Detalhes"
        'empresa': r[5],
        'online': r[6], 'convenio': r[7],
        't1': r[8], 'tret': r[9], 'valor': r[10],
        'idade': r[11], 'crm': r[12], 'rqe': r[13],
        'cpf': r[14], 'cel': r[15], 'email': r[16], 'insta': r[17],
    })
print(f'Corpo Clínico: {len(profs)} profissionais')

# ===== 2. MAPA DE SALA =====
wb2 = openpyxl.load_workbook(PLAN_MAPA, read_only=True, data_only=True)
ws2 = wb2['Mapa']
mrows = list(ws2.iter_rows(min_row=4, max_row=200, values_only=True))
salas = [str(c).replace('\n', ' ').strip() for c in mrows[0][2:12]]

DIAS = {'SEGUNDA-FEIRA':1,'TERÇA-FEIRA':2,'QUARTA-FEIRA':3,'QUINTA-FEIRA':4,
        'SEXTA-FEIRA':5,'SÁBADO':6}
grade = []   # (dia, hora, sala_idx, texto)
dia = None
for r in mrows[1:]:
    d = (str(r[0]).strip().upper() if r[0] else '')
    if d in DIAS: dia = DIAS[d]
    hora = r[1]
    for i, c in enumerate(r[2:12]):
        if c and str(c).strip() and dia:
            grade.append((dia, hora, i, str(c).strip()))
print(f'Mapa: {len(salas)} salas, {len(grade)} blocos de escala')

# ===== 3. RECONCILIAÇÃO =====
# extrai CRM do texto do mapa pra casar com o Corpo Clínico
def crm_do_texto(t):
    m = re.search(r'CRM\s*(\d+)', t)
    return m.group(1) if m else None

# Remove prefixo Dr./Dra. — Dra? casa ambos (regex antigo deixava "a." sobrando)
def sem_titulo(nome):
    return re.sub(r'^dra?\.?\s*', '', str(nome or ''), flags=re.I).strip()

def primeiro_ultimo(nome):
    p = sem_titulo(nome).split()
    return (norm(p[0]), norm(p[-1])) if p else ('', '')

def nomes_casam(a, b):
    """a, b = (primeiro, último) normalizados. Tolera acento, grafia e
    abreviação (Carol/Carolina, Klepacs/Klepacz)."""
    pf = a[0] == b[0] or a[0].startswith(b[0]) or b[0].startswith(a[0])
    pu = a[1] == b[1] or a[1][:5] == b[1][:5]
    return pf and pu

# nomes que aparecem no Mapa mas precisam virar ficha
nomes_mapa = {}
for dia, hora, si, txt in grade:
    nome_limpo = re.split(r'\s*-\s*', txt)[0].strip()
    nomes_mapa.setdefault(nome_limpo, txt)

# casa cada nome do mapa com um índice de profs (ou marca como novo)
prof_por_crm = {str(p['crm']).split('.')[0]: idx for idx, p in enumerate(profs) if p['crm']}
def acha_prof(txt):
    crm = crm_do_texto(txt)
    if crm and crm in prof_por_crm:
        return prof_por_crm[crm]
    nome = re.split(r'\s*-\s*', txt)[0].strip()
    alvo = primeiro_ultimo(nome)
    for idx, p in enumerate(profs):
        if nomes_casam(alvo, primeiro_ultimo(p['nome'])):
            return idx
    return None

# profissionais só no Mapa → criar ficha parcial
extras = {}  # nome_limpo -> dict
for nome_limpo, txt in nomes_mapa.items():
    if acha_prof(txt) is None and nome_limpo not in extras:
        partes = re.split(r'\s*-\s*', txt)
        espec = partes[1].strip() if len(partes) > 1 else None
        extras[nome_limpo] = {
            'nome': sem_titulo(nome_limpo),
            'especialidade': espec, 'observacoes': '[só no Mapa de Sala — ficha a completar]',
            'empresa': 'Endovascular', 'online': None, 'convenio': None,
            't1': None, 'tret': None, 'valor': None, 'idade': None,
            'crm': None, 'rqe': None, 'cpf': None, 'cel': None, 'email': None, 'insta': None,
        }
todos = profs + list(extras.values())
print(f'Total: {len(profs)} do Corpo Clínico + {len(extras)} só do Mapa = {len(todos)}')

# índice nome->id (1-based)
def id_do_texto(txt):
    idx = acha_prof(txt)
    if idx is not None: return idx + 1
    nome_limpo = re.split(r'\s*-\s*', txt)[0].strip()
    chaves = list(extras.keys())
    if nome_limpo in chaves:
        return len(profs) + chaves.index(nome_limpo) + 1
    return None

# ===== 4. GERA SQL =====
L = []
L.append('-- SEED — Portal Corpo Clínico (gerado de Planilha Geral + Mapa de Sala)\n')
L.append('-- 1) Salas')
L.append('INSERT INTO cc_salas (id, nome, tipo, ordem) VALUES')
linhas = []
for i, s in enumerate(salas):
    tl = s.lower()
    tipo = ('fotona' if 'fotona' in tl else 'spa' if tl == 'spa'
            else 'procedimento' if 'procedimento' in tl else 'consultorio')
    linhas.append(f"  ({i+1}, {esc(s)}, '{tipo}', {i+1})")
L.append(',\n'.join(linhas) + ';\n')

L.append('-- 2) Profissionais')
L.append('INSERT INTO cc_profissionais (id, nome, especialidade, empresa, atende_online,')
L.append('  atende_convenio, tempo_primeira_vez, tempo_retorno, valor_consulta, faixa_idade,')
L.append('  conselho, conselho_numero, rqe, cpf, celular, email, instagram, observacoes) VALUES')
linhas = []
for i, p in enumerate(todos):
    consel = conselho_de(p['especialidade'])
    cnum = str(p['crm']).split('.')[0] if p['crm'] else None
    linhas.append(
        f"  ({i+1}, {esc(p['nome'])}, {esc(p['especialidade'])}, {esc(p['empresa'])}, "
        f"{boolsql(p['online']) if p['online'] is not None else 'NULL'}, "
        f"{boolsql(p['convenio']) if p['convenio'] is not None else 'NULL'}, "
        f"{esc(p['t1'])}, {esc(p['tret'])}, {num(p['valor'])}, {esc(p['idade'])}, "
        f"{esc(consel)}, {esc(cnum)}, {esc(p['rqe'])}, {esc(p['cpf'])}, "
        f"{esc(limpa_tel(p['cel']))}, {esc(p['email'])}, {esc(p['insta'])}, {esc(p['observacoes'])})")
L.append(',\n'.join(linhas) + ';\n')

L.append('-- 3) Escala (grade do Mapa de Sala). hora_fim fica NULL — Agendamento confirma.')
L.append('INSERT INTO cc_escala (profissional_id, sala_id, dia_semana, hora_inicio) VALUES')
linhas = []
sem_match = []
for dia, hora, si, txt in grade:
    pid = id_do_texto(txt)
    if pid is None:
        sem_match.append(txt); continue
    hs = hora.strftime('%H:%M') if hasattr(hora, 'strftime') else str(hora)
    linhas.append(f"  ({pid}, {si+1}, {dia}, '{hs}')")
L.append(',\n'.join(linhas) + ';\n')

L.append('-- Reseta as sequences pros próximos inserts via portal')
L.append("SELECT setval('cc_salas_id_seq',         (SELECT MAX(id) FROM cc_salas));")
L.append("SELECT setval('cc_profissionais_id_seq', (SELECT MAX(id) FROM cc_profissionais));")
L.append("SELECT setval('cc_escala_id_seq',        (SELECT MAX(id) FROM cc_escala));")

with open(OUT, 'w', encoding='utf-8', newline='') as f:
    f.write('\n'.join(L) + '\n')

print(f'\nSeed gerado: {OUT}')
if sem_match:
    print(f'AVISO — {len(sem_match)} blocos de escala sem profissional casado:')
    for t in sem_match: print(f'  - {t}')

# -*- coding: utf-8 -*-
"""
_honorarios_catalogo.py — tradução procedimento -> categoria, num lugar só.

O catálogo tem duas origens:

  1. A aba "Apoio" do Excel de fechamento (184 linhas, 177 procedimentos únicos).
     É o legado — continua sendo lida enquanto o Excel existir.

  2. EXTRAS, abaixo: procedimentos que apareceram depois e foram classificados
     pelo Thiago ao resolver a fila de exceções.

Tanto o motor quanto o gerador do seed usam esta função, para que o catálogo do
código e o do Supabase nunca divirjam.
"""
from __future__ import annotations
import unicodedata

PLANILHA = r"G:\Drives compartilhados\Endovascular SP\2. Financeiro\4. Honorários médicos\Fechamento - Endovascular SP.xlsx"

# Grafias que o Excel deixou duplicadas — a da direita é a canônica.
CANONICA = {
    "exames de imagem": "Exames de imagem",
    "medicacao injetavel": "Medicação injetável",
}

# Classificados pelo Thiago em 03/08/2026, ao resolver a fila de Julho.
EXTRAS = {
    "Cirurgia - Tenotomia":                                                   "Cirurgia - Hospital",
    "Preenchimento (Por seringa) - Tabela Diretoria":                         "Procedimentos",
    "Exossomos - Terapia Regenerativa":                                       "Procedimentos",
    "Hybrius EVO - Sessão individual (1 área)":                               "Laser (clínica)",
    "Exérese e sutura simples de pequenas lesões (por grupo de até 5 lesões)": "Procedimentos",
    "Taxa compacta de sala de pequenas cirurgias":                            "Cirurgia - Clínica",

    # Classificados em 05/08/2026, ao migrar a Produtividade para o relatório #560.
    # A maioria já existia no catálogo com o nome levemente diferente — o SVN tem
    # o mesmo procedimento cadastrado com e sem o prefixo "Cirurgia -", ou com
    # variação de grafia.
    "Cirurgia - Aneurisma de axilar, femoral, poplítea":                      "Cirurgia - Hospital",
    "Varizes - tratamento cirúrgico por radiofrequência de dois membros":     "Cirurgia - Hospital",
    "Cirurgia - Restauração venosa com pontes nos membros":                   "Cirurgia - Hospital",
    "Embolização de malformação vascular - por vaso":                         "Cirurgia - Hospital",
    "Retirada cirúrgica de cateter de longa permanência para NPP, QT ou para Hemodepuração": "Cirurgia - Hospital",
    "Fotona Dores e Inflamações - ConfortLase":                               "Fotona",
    "Onicomicoses":                                                           "Fotona",
    "Noripurum (2 Ampolas)":                                                  "Medicação injetável",
    "Vitamina D 50.000UI":                                                    "Medicação injetável",
    "T-SCULPTOR - 8 Sessões (4 áreas)":                                       "T-Sculptor",
    "Consulta Nutricionista":                                                 "Consultas",
    "Consulta Pós-Operatória":                                                "Consultas",
    "Doppler colorido de veia cava superior ou inferior (Cardio)":            "Exames de imagem",
    # Mesmo exame que "Avaliação da composição corporal por bioimpedanciometria",
    # cadastrado com nome curto. Aparece nas duas empresas.
    "Bioimpedância":                                                          "Exames gerais",
    "Bioimpedanciometria (ambulatorial) exame":                               "Exames gerais",
    # Sem equivalente no catálogo; classificados pelo tipo de procedimento.
    "Revascularização de Aorta Bi-Femoral (Convencional)":                    "Cirurgia - Hospital",
    "Colocação de stent renal":                                               "Cirurgia - Hospital",
    "Drenagem Linfática Manual":                                              "Fisioterapia",
}

# ---------------------------------------------------------------------------
# Procedimentos AMBÍGUOS — nunca entram no catálogo
# ---------------------------------------------------------------------------
# Cadastro legado: o campo Procedimento traz só "Laser", que pode ser tanto
# Laser Transdérmico (categoria "Laser (clínica)") quanto fibra de laser usada em
# cirurgia (categoria "Cirurgia - Hospital"). O NOME não permite decidir — a
# semelhança com a categoria "Laser (clínica)" é coincidência e cadastrar por ela
# jogaria linhas cirúrgicas de 80-90% para 60%, de forma silenciosa.
#
# Só o contexto da OS resolve: ver quais outros procedimentos foram lançados na
# mesma OS. Por isso ficam SEMPRE fora do catálogo e SEMPRE caem na fila.
AMBIGUOS = {"laser", "laser - pacote"}

# ---------------------------------------------------------------------------
# Resolução por linha: (Nº OS, chave do procedimento) -> categoria
# ---------------------------------------------------------------------------
# É o que a fila de exceções produz quando o nome do procedimento não basta.
# Decidido pelo Thiago em 03/08/2026, a partir do cruzamento por OS.
#
# Resolvidos em 05/08/2026 cruzando com `honorarios_lancamentos`: a mesma OS já
# aparece no fechamento com o procedimento de nome completo e categoria atribuída.
# Essa fonte resolveu 20 de 20 OS, contra 3 de 20 do cruzamento por vizinhança —
# é o caminho a tentar primeiro nos próximos casos.
OVERRIDES_POR_OS = {
    # Onde a mesma OS traz "Endolaser Cirúrgico" (+ varizes) -> é a fibra de laser
    ("13055855", "laser"):          "Cirurgia - Hospital",  # Karina Geraldini / Igor
    ("13647510", "laser"):          "Cirurgia - Hospital",  # Cristiane de Araújo / Igor
    ("13750002", "laser"):          "Cirurgia - Hospital",  # Paulo Emilio / Manoel
    ("13892404", "laser"):          "Cirurgia - Hospital",  # Laurinda Yamanishi / Manoel
    ("13225365", "laser - pacote"): "Cirurgia - Hospital",  # Débora de Mattos / Igor

    # Onde a mesma OS traz "Laser Transdérmico" -> é laser de clínica
    ("13694521", "laser"): "Laser (clínica)",  # Mélanie Isabelle Smuga / Igor
    ("13735169", "laser"): "Laser (clínica)",  # Tatiana Armonas Seide
    ("13748605", "laser"): "Laser (clínica)",  # Sandra Regina Leite / Manoel
    ("13760719", "laser"): "Laser (clínica)",  # Débora Dutra / Igor
    ("13760729", "laser"): "Laser (clínica)",  # Débora Dutra / Igor
    ("13790902", "laser"): "Laser (clínica)",  # Ana Maria de Jesus Frade / Manoel
    ("13820187", "laser"): "Laser (clínica)",  # Terezinha de Freitas / Manoel
    ("13820636", "laser"): "Laser (clínica)",  # Simone de Souza Silva / Manoel
    ("13834547", "laser"): "Laser (clínica)",  # Maria Aguida de Lima / Andrea
    ("13834861", "laser"): "Laser (clínica)",  # Mariana Bonavita / Andrea
    ("13858952", "laser"): "Laser (clínica)",  # Josefa Hortencia / Manoel
    ("13887021", "laser"): "Laser (clínica)",  # Luciana Steiner / João Fukuda
    ("13887077", "laser"): "Laser (clínica)",  # Marcia Weiser / João Fukuda
    ("13903804", "laser"): "Laser (clínica)",  # Juliana Pancev Danez / Clara
    ("13904566", "laser"): "Laser (clínica)",  # Mélanie Isabelle Smuga / Igor
    ("13937525", "laser"): "Laser (clínica)",  # Arturania Diniz / Andrea
    ("13938040", "laser"): "Laser (clínica)",  # Juliane Gomes de Paula / Clara
}


def chave(s) -> str:
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return " ".join(s.lower().split())


def carregar() -> dict:
    """{chave_normalizada: categoria} — lido do Supabase.

    Desde 07/08/2026 o catálogo vive na tabela `honorarios_procedimentos`, não
    mais na aba "Apoio" do Excel. Foi o último fio que prendia o fechamento à
    planilha: sem isto, apagar o arquivo do Drive quebrava a geração do mês.

    EXTRAS continua no código como fila de espera: procedimentos classificados
    aqui e ainda não semeados no banco. O seed (`--gravar`) os leva para lá; a
    partir daí eles voltam por esta função e a lista pode ser esvaziada.
    """
    import _honorarios_db as DB
    cat = {r["chave"]: r["categoria"]
           for r in DB.buscar("honorarios_procedimentos", "chave,categoria")}
    if not cat:
        raise SystemExit("ABORTADO: honorarios_procedimentos está vazia. "
                         "Rode _tools/_honorarios_seed_procedimentos.py --gravar")
    # EXTRAS ainda não semeados entram por cima, para o mês não travar por causa
    # de um seed esquecido. O seed é que faz isso virar permanente.
    for proc, c in EXTRAS.items():
        cat.setdefault(chave(proc), CANONICA.get(chave(c), c))
    return cat


def carregar_do_excel(planilha: str = PLANILHA) -> dict:
    """A versão antiga, lendo a aba "Apoio". Só o seed usa — é a ponte que leva
    o catálogo da planilha para o banco. Nenhum caminho do fechamento passa
    mais por aqui."""
    import openpyxl
    cat = {}
    wb = openpyxl.load_workbook(planilha, data_only=True, read_only=True)
    for r in wb["Apoio"].iter_rows(min_row=3, max_row=500, min_col=2, max_col=3, values_only=True):
        if r[0] and r[1]:
            cat.setdefault(chave(r[0]), CANONICA.get(chave(r[1]), str(r[1]).strip()))
    wb.close()
    for proc, c in EXTRAS.items():
        cat[chave(proc)] = CANONICA.get(chave(c), c)
    return cat


def categoria_de(procedimento, os_numero, catalogo: dict):
    """(categoria, origem) — ou (None, motivo) quando precisa de decisão humana.

    Ordem: override da linha vence o catálogo; procedimento ambíguo nunca é
    resolvido pelo nome."""
    k = chave(procedimento)
    over = OVERRIDES_POR_OS.get((str(os_numero).strip(), k))
    if over:
        return over, "categoria definida por OS"
    if k in AMBIGUOS:
        return None, (f"Procedimento '{procedimento}' é ambíguo: o nome não diz a "
                      "categoria. Ver os outros procedimentos da mesma OS.")
    cat = catalogo.get(k)
    if cat:
        return cat, None
    return None, f"Procedimento '{procedimento}' não está no catálogo"


def itens(planilha: str = PLANILHA) -> list:
    """[(chave, procedimento_original, categoria)] — para gerar o seed SQL."""
    import openpyxl
    out, vistos = [], set()
    wb = openpyxl.load_workbook(planilha, data_only=True, read_only=True)
    linhas = [(r[0], r[1]) for r in wb["Apoio"].iter_rows(min_row=3, max_row=500,
                                                          min_col=2, max_col=3,
                                                          values_only=True) if r[0] and r[1]]
    wb.close()
    for proc, c in linhas + list(EXTRAS.items()):
        k = chave(proc)
        if k in vistos:
            continue
        vistos.add(k)
        out.append((k, str(proc).strip(), CANONICA.get(chave(c), str(c).strip())))
    return out

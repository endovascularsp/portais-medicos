"""Gera SQL de cadastro em massa dos funcionários a partir do Excel.

Lê User_Download_*.xlsx e gera:
  1. INSERT dos departamentos faltantes em compras_departamentos
  2. INSERT dos 23 funcionários em users (ON CONFLICT email DO NOTHING)

Regras:
  - Departamento 'Gestão'  → role=admin,    cards=['gestor']
  - Demais departamentos   → role=recepcao, cards=['compras_solicitante']
  - 'Adminstrativo' (typo no Excel) corrigido pra 'Administrativo'
  - departamento_id resolvido via subquery (não hardcoda id)
  - ON CONFLICT (email) DO NOTHING → não sobrescreve quem já existe

CONFIG: ajuste CARDS_FORMAT conforme o tipo da coluna users.cards:
  - 'jsonb'  → cards viram '["x"]'::jsonb
  - 'array'  → cards viram ARRAY['x']
"""
import openpyxl, sys

CARDS_FORMAT = 'array'   # users.cards é text[] (confirmado: data_type ARRAY)

XLSX = 'C:/Users/thiag/Downloads/User_Download_20052026_231213_File.xlsx'
OUT  = 'C:/Users/thiag/Documents/portais-medicos/compras/_setup/06_seed_funcionarios.sql'

# Correções de nome de departamento (Excel → canônico)
DEPT_FIX = {'Adminstrativo': 'Administrativo'}

def fmt_cards(lista):
    if CARDS_FORMAT == 'jsonb':
        import json
        return "'" + json.dumps(lista) + "'::jsonb"
    else:  # array
        return 'ARRAY[' + ','.join(f"'{c}'" for c in lista) + ']'

def esc(s):
    return str(s or '').replace("'", "''").strip()

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[2]]

# Departamentos únicos (corrigidos)
deptos = set()
for fn, ln, email, dep, st in rows:
    dep = DEPT_FIX.get((dep or '').strip(), (dep or '').strip())
    if dep: deptos.add(dep)

linhas = []
linhas.append('-- =====================================================================')
linhas.append('-- SEED FUNCIONÁRIOS — cadastro em massa (23 pessoas)')
linhas.append('-- Gerado de User_Download_20052026_231213_File.xlsx')
linhas.append('-- =====================================================================\n')

linhas.append('-- 1) Departamentos faltantes (ON CONFLICT evita duplicar os 6 do seed)')
linhas.append('INSERT INTO compras_departamentos (nome) VALUES')
linhas.append(',\n'.join(f"  ('{esc(d)}')" for d in sorted(deptos)))
linhas.append('ON CONFLICT (nome) DO NOTHING;\n')

linhas.append('-- 2) Funcionários. ON CONFLICT (email) DO NOTHING preserva quem já existe.')
linhas.append('INSERT INTO users (email, name, role, cards, departamento_id) VALUES')

vals = []
for fn, ln, email, dep, st in rows:
    dep = DEPT_FIX.get((dep or '').strip(), (dep or '').strip())
    nome = esc(f"{fn or ''} {ln or ''}".strip())
    email_l = esc(email).lower()
    if dep == 'Gestão':
        role, cards = 'admin', ['gestor']
    else:
        role, cards = 'recepcao', ['compras_solicitante']
    dep_sub = f"(SELECT id FROM compras_departamentos WHERE nome='{esc(dep)}')"
    vals.append(f"  ('{email_l}', '{nome}', '{role}', {fmt_cards(cards)}, {dep_sub})")

linhas.append(',\n'.join(vals))
linhas.append('ON CONFLICT (email) DO NOTHING;')

sql = '\n'.join(linhas) + '\n'
with open(OUT, 'w', encoding='utf-8', newline='') as f:
    f.write(sql)

# Resumo
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
print(f"SQL gerado: {OUT}")
print(f"  Formato cards: {CARDS_FORMAT}")
print(f"  Departamentos: {len(deptos)} ({', '.join(sorted(deptos))})")
gestao = sum(1 for r in rows if DEPT_FIX.get((r[3] or '').strip(),(r[3] or '').strip())=='Gestão')
print(f"  Funcionários: {len(rows)} ({gestao} Gestão=admin, {len(rows)-gestao} demais=recepcao)")
